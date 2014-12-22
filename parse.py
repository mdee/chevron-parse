# -*- coding: utf-8 -*-
from datetime import datetime
import os
import re
from enum import Enum
import sys
from os.path import isdir
from openpyxl import load_workbook, Workbook
from openpyxl.exceptions import InvalidFileException
from openpyxl.styles import Font, PatternFill, Style, Color

Location = Enum('Location', 'Indoor Outdoor')
Gas = Enum('Gas', 'UNLEADED PLUS SUPREME')
Carwash = Enum('Carwash', 'Regular Deluxe Super')
Tender = Enum('Tender', 'Cash Credit Debit')
DIRECTORY_ANALYZED_FILENAME = 'ALREADY_ANALYZED'
FINALIZED_TRANSACTION_REGEX = re.compile(r'CUSTOMER\sTRANSACTION\s+(?P<txn_id>[0-9]+)\s+Finalized')
INDOOR_OUTDOOR_REGEX = re.compile(r'(?P<location>(Indoor|Outdoor)+)\s+tmnl(\s+)?:\s+(?P<terminal>[0-9]+)')
USER_SESSION_REGEX = re.compile(r'User\s+Session:\s+[0-9]+')
INITIAL_FUEL_PREPAY_REGEX = re.compile(r'Fuel\s+Prepay\s+Ref#(?P<ref_num>[0-9]+)\s+Pump\s+(?P<pump_num>[0-9]+)')
FINAL_FUEL_PREPAY_REGEX = re.compile(r'Original\s+Fuel\s+Prepay\s+Ref#(?P<ref_num>[0-9]+)')
FUEL_PREPAY_AMOUNT_REGEX = re.compile(r'FUEL\s+PREPAY\s+(?P<dollars>[0-9]+)\.(?P<cents>[0-9]+)')
TOTAL_DUE_REGEX = re.compile(r'TOTAL\s+DUE\s+[0-9]+\.[0-9]+')
BALANCE_DUE_REGEX = re.compile(r'BALANCE\s+DUE\s+[0-9]+\.[0-9]+')
INDOOR_TENDER_TYPE_REGEX = re.compile(r'(?P<tender>\w+)\s+.*')
DATE_TIME_REGEX = re.compile(r'(?P<month>[0-9]+)/(?P<day>[0-9]+)/(?P<yr>[0-9]+)\s+(?P<hr>[0-9]+):(?P<min>[0-9]+):(?P<sec>[0-9]+)')
FUEL_TYPE_REGEX = re.compile(r'\s+(?P<gas_type>(PLUS|UNLEADED|SUPREME)+)\s+PUR(E)?\s+(?P<dollars>[0-9]+)\.(?P<cents>[0-9]+)')
FUEL_VOLUME_REGEX = re.compile(r'\s+Vol\s+(?P<galls>[0-9]+)\.(?P<galls_dec>[0-9]+)@\s+(?P<price>[0-9]+)\.(?P<price_dec>[0-9]+)')
OUTDOOR_TENDER_TYPE_REGEX = re.compile(r'(?P<tender>(Credit|Debit)+)\s+Card\s+[0-9]+.[0-9]+')
VOID_TRANSACTION_REGEX = re.compile(r'\s+\*Void\*\s+')
GAS_CELL_LABELS = ['gallons < 5', '5 < gallons < 10', '10 < gallons < 14', '14 < gallons < 19', 'gallons > 19', 'Unleaded', 'Plus', 'Supreme', 'Indoor', 'Outdoor', 'Credit card', 'Debit card', 'Cash', 'Total carwash', 'Regular', 'Deluxe', 'Super', 'Regular indoor', 'Regular outdoor', 'Deluxe indoor', 'Deluxe outdoor', 'Super indoor', 'Super outdoor']
CARWASH_REGEX = re.compile(r'\s+CAR\s+WASH\s+(?P<type>(SUP|DEL|\-\s+W))\s+(\-)?(?P<dollars>[0-9]+)\.(?P<cents>[0-9]+)')


class Txn(object):
    def __init__(self, id, date, time, amount, location, tender):
        """"""
        self.id = id
        self.date = date
        self.time = time
        self.amount = amount
        self.location = location
        self.tender = tender


class GasTxn(Txn):
    def __init__(self, id, date, time, amount, location, tender, volume, gas_type, pump_num, indoor_prepay=False, reference_num=None, price=None):
        """"""
        Txn.__init__(self, id, date, time, amount, location, tender)
        self.volume = volume if not volume else float(volume)
        self.gas_type = gas_type
        self.pump_num = pump_num if not pump_num else int(pump_num)
        self.indoor_prepay = indoor_prepay
        self.reference_num = reference_num
        self.price = price
        self.carwash_txn = None


class CarWashTxn(Txn):
    """"""
    def __init__(self, id, date, time, amount, location, carwash_type, tender):
        Txn.__init__(self, id, date, time, amount, location, tender)
        if carwash_type == '- W':
            self.carwash_type = Carwash.Regular.name
        elif carwash_type == 'DEL':
            self.carwash_type = Carwash.Deluxe.name
        else:
            self.carwash_type = Carwash.Super.name


class DayAnalyzer(object):
    def __init__(self, gas_txns, date, index, carwash_txns):
        self.gas_txns = gas_txns
        self.date = date
        self.index = index
        self.carwash_txns = carwash_txns

    def add_gas_txns_to_worksheet(self, ws):
        """"""
        bold_font = Font(bold=True)
        label_fill = PatternFill(fill_type='solid', start_color=Color('00EEEEEE'))
        label_cell_style = Style(font=bold_font, fill=label_fill)
        label_cell = ws.cell(get_column_letter_for_column_number(1, self.index))
        label_cell.style = label_cell_style
        label_cell.value = self.date.strftime('%m/%d')
        # ['gallons < 5', '5 < gallons < 10', '10 < gallons < 14', '14 < gallons < 19', 'gallons > 19', 'Unleaded', 'Plus', 'Supreme', 'Indoor', 'Outdoor', 'Credit card', 'Debit card', 'Cash', 'Total carwash', 'Regular', 'Deluxe', 'Super', 'Regular indoor', 'Regular outdoor', 'Deluxe indoor', 'Deluxe outdoor', 'Super indoor', 'Super outdoor']
        row_index = 2
        lt5_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        lt5_cell.value = self.get_vol_less_than(less_than=5.0)
        row_index += 1
        btw_5_10_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        btw_5_10_cell.value = self.get_vol_btw(greater_than=5.0, less_than=10.0)
        row_index += 1
        btw_10_14 = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        btw_10_14.value = self.get_vol_btw(greater_than=10.0, less_than=14.0)
        row_index += 1
        btw_14_19 = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        btw_14_19.value = self.get_vol_btw(greater_than=14.0, less_than=19.0)
        row_index += 1
        gt_20 = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        gt_20.value = self.get_vol_greater_than(greater_than=19.0)
        row_index += 1
        unleaded_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        unleaded_cell.value = self.get_count_with_gas_type(Gas.UNLEADED.name)
        row_index += 1
        plus_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        plus_cell.value = self.get_count_with_gas_type(Gas.PLUS.name)
        row_index += 1
        supreme_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        supreme_cell.value = self.get_count_with_gas_type(Gas.SUPREME.name)
        row_index += 1
        indoor_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        indoor_cell.value = self.get_location_count(Location.Indoor.name)
        row_index += 1
        outdoor_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        outdoor_cell.value = self.get_location_count(Location.Outdoor.name)
        row_index += 1
        credit_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        credit_cell.value = self.get_tender_count(Tender.Credit.name)
        row_index += 1
        debit_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        debit_cell.value = self.get_tender_count(Tender.Debit.name)
        row_index += 1
        cash_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        cash_cell.value = self.get_tender_count(Tender.Cash.name)
        regular_indoor = self.get_wash_type_count_for_location(Location.Indoor.name, Carwash.Regular.name)
        regular_outdoor = self.get_wash_type_count_for_location(Location.Outdoor.name, Carwash.Regular.name)
        deluxe_indoor = self.get_wash_type_count_for_location(Location.Indoor.name, Carwash.Deluxe.name)
        deluxe_outdoor = self.get_wash_type_count_for_location(Location.Outdoor.name, Carwash.Deluxe.name)
        super_indoor = self.get_wash_type_count_for_location(Location.Indoor.name, Carwash.Super.name)
        super_outdoor = self.get_wash_type_count_for_location(Location.Outdoor.name, Carwash.Super.name)
        row_index += 1
        total_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        total_cell.value = regular_indoor + regular_outdoor + deluxe_indoor + deluxe_outdoor + super_indoor + super_outdoor
        row_index += 1
        reg_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        reg_cell.value = regular_indoor + regular_outdoor
        row_index += 1
        del_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        del_cell.value = deluxe_indoor + deluxe_outdoor
        row_index += 1
        sup_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        sup_cell.value = super_indoor + super_outdoor
        row_index += 1
        regi_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        regi_cell.value = regular_indoor
        row_index += 1
        rego_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        rego_cell.value = regular_outdoor
        row_index += 1
        deli_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        deli_cell.value = deluxe_indoor
        row_index += 1
        delo_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        delo_cell.value = deluxe_outdoor
        row_index += 1
        supi_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        supi_cell.value = super_indoor
        row_index += 1
        supo_cell = ws.cell(get_column_letter_for_column_number(row_index, self.index))
        supo_cell.value = super_outdoor

    def get_pump_count(self, pump):
        count = 0
        for t in self.gas_txns:
            if t.pump_num == pump:
                count += 1
        return count

    def get_wash_type_count_for_location(self, location, wash_type):
        """"""
        count = 0
        for g in self.gas_txns:
            if g.carwash_txn and g.carwash_txn.carwash_type == wash_type and g.location == location:
                count += 1
        for c in self.carwash_txns:
            if c.carwash_type == wash_type and c.location == location:
                count += 1
        return count

    def get_tender_count(self, tender):
        count = 0
        for t in self.gas_txns:
            if t.tender == tender:
                count += 1
        return count

    def get_location_count(self, location):
        count = 0
        for t in self.gas_txns:
            if t.location == location:
                count += 1
        return count

    def get_count_with_gas_type(self, gas_type):
        count = 0
        for t in self.gas_txns:
            if t.gas_type == gas_type:
                count += 1
        return count

    def get_vol_btw(self, greater_than, less_than):
        count = 0
        for t in self.gas_txns:
            if greater_than < t.volume < less_than:
                count += 1
        return count

    def get_vol_greater_than(self, greater_than):
        count = 0
        for t in self.gas_txns:
            if t.volume > greater_than:
                count += 1
        return count

    def get_vol_less_than(self, less_than):
        """"""
        count = 0
        for t in self.gas_txns:
            if t.volume < less_than:
                count += 1
        return count


def touch(fname, times=None):
    """"""
    with open(fname, 'a'):
        os.utime(fname, times)


def get_month_directories_to_analyze(dir_paths):
    """"""
    for p in dir_paths:
        if DIRECTORY_ANALYZED_FILENAME not in os.listdir(p):
            yield p


def mark_directory_as_analyzed(dir_path):
    """"""
    touch(fname=os.path.join(dir_path, DIRECTORY_ANALYZED_FILENAME))


def handle_initial_gas_txn(fuel_prepay_ref_match, txn_line_offset, line_offsets, txn_line, f, txn_id, date, time):
    """"""
    # Check to make sure there isn't a void. If there is, the fuel_prepay_ref_match needs to be updated
    void_offset = 0
    void_found = False
    while not void_found:
        void_offset += 1
        f.seek(line_offsets[txn_line + void_offset])
        l = f.readline()
        if re.match(VOID_TRANSACTION_REGEX, l):
            void_found = True
        elif re.match(TOTAL_DUE_REGEX, l):
            break
    # If there was a void transaction, advance until you find the new fuel prepay, or if total due is hit you're fine
    # Otherwise, seek backwards
    if void_found:
        fuel_prepay_found = False
        while not fuel_prepay_found:
            void_offset += 1
            f.seek(line_offsets[txn_line + void_offset])
            l = f.readline()
            new_prepay_match = re.match(INITIAL_FUEL_PREPAY_REGEX, l)
            if new_prepay_match:
                fuel_prepay_found = True
                fuel_prepay_ref_match = new_prepay_match
            elif re.match(TOTAL_DUE_REGEX, l):
                break
    else:
        f.seek(line_offsets[txn_line + txn_line_offset])
    # We've got a prepay, create a new gas txn
    ref_num = fuel_prepay_ref_match.group('ref_num')
    pump_num = fuel_prepay_ref_match.group('pump_num')
    # Seek to the next line to get the amount
    txn_line_offset += 1
    f.seek(line_offsets[txn_line + txn_line_offset])
    f_pre_amt_line = f.readline()
    amt_match = re.match(FUEL_PREPAY_AMOUNT_REGEX, f_pre_amt_line)
    amount = float('{0}.{1}'.format(amt_match.group('dollars'), amt_match.group('cents')))
    # Seek until you reach the TOTAL DUE line
    total_due_line_found = False
    while not total_due_line_found:
        txn_line_offset += 1
        f.seek(line_offsets[txn_line + txn_line_offset])
        l = f.readline()
        if re.match(TOTAL_DUE_REGEX, l):
            total_due_line_found = True
    # Seek ahead one line, if it's BALANCE DUE then it's cash and you're good
    txn_line_offset += 1
    f.seek(line_offsets[txn_line + txn_line_offset])
    next_line = f.readline()
    if re.match(BALANCE_DUE_REGEX, next_line):
        tender = Tender.Cash.name
    else:
        # Seek ahead one more line, and pull the word out
        txn_line_offset += 1
        f.seek(line_offsets[txn_line + txn_line_offset])
        t_type_line = f.readline()
        tender_type_match = re.match(INDOOR_TENDER_TYPE_REGEX, t_type_line)
        if tender_type_match:
            tender = tender_type_match.group('tender')
        else:
            tender = Tender.Cash.name
    prepay_txn = GasTxn(id=txn_id, date=date, time=time, amount=amount, location=Location.Indoor.name, tender=tender, volume=None, gas_type=None, pump_num=pump_num, indoor_prepay=True, reference_num=ref_num)
    carwash_txn = scan_for_carwash_from_line(txn_line, line_offsets, f)
    if carwash_txn:
        prepay_txn.carwash_txn = carwash_txn
    return True, prepay_txn


def handle_final_gas_txn(original_fuel_prepay_match, txn_line_offset, line_offsets, txn_line, f, txn_id, date, time):
    """"""
    ref_num = original_fuel_prepay_match.group('ref_num')
    # Seek ahead to get the gas type
    fuel_type_found, fuel_type_match = False, None
    while not fuel_type_found:
        txn_line_offset += 1
        f.seek(line_offsets[txn_line + txn_line_offset])
        t_line = f.readline()
        fuel_type_match = re.match(FUEL_TYPE_REGEX, t_line)
        if fuel_type_match:
            fuel_type_found = True
    if not fuel_type_match:
        return False, None
    fuel_type = fuel_type_match.group('gas_type')
    # Seek forward two lines to get the volume
    txn_line_offset += 2
    f.seek(line_offsets[txn_line + txn_line_offset])
    vol_line = f.readline()
    vol_match = re.match(FUEL_VOLUME_REGEX, vol_line)
    fuel_volume = '{0}.{1}'.format(vol_match.group('galls'), vol_match.group('galls_dec'))
    price = '{0}.{1}'.format(vol_match.group('price'), vol_match.group('price_dec'))
    final_txn = GasTxn(id=txn_id, date=date, time=time, amount=None, volume=fuel_volume, location=Location.Indoor.name, tender=None, gas_type=fuel_type, pump_num=None, indoor_prepay=False, reference_num=ref_num, price=price)
    return True, final_txn


def handle_outdoor_gas_txn(txn_id, date, time, pump_num, txn_line_offset, line_offsets, txn_line, f):
    """"""
    # Advance 2 lines to get fuel type
    txn_line_offset += 3
    f.seek(line_offsets[txn_line + txn_line_offset])
    f_type_line = f.readline()
    fuel_type_match = re.match(FUEL_TYPE_REGEX, f_type_line)
    fuel_type = fuel_type_match.group('gas_type')
    amount = '{0}.{1}'.format(fuel_type_match.group('dollars'), fuel_type_match.group('cents'))
    # Advance 2 lines to get volume
    txn_line_offset += 2
    f.seek(line_offsets[txn_line + txn_line_offset])
    vol_line = f.readline()
    vol_match = re.match(FUEL_VOLUME_REGEX, vol_line)
    fuel_volume = '{0}.{1}'.format(vol_match.group('galls'), vol_match.group('galls_dec'))
    price = '{0}.{1}'.format(vol_match.group('price'), vol_match.group('price_dec'))
    tender_type_found, tender = False, None
    while not tender_type_found:
        txn_line_offset += 1
        f.seek(line_offsets[txn_line + txn_line_offset])
        t_line = f.readline()
        tender_match = re.match(OUTDOOR_TENDER_TYPE_REGEX, t_line)
        if tender_match:
            tender_type_found = True
            tender = tender_match.group('tender')
    outdoor_txn = GasTxn(id=txn_id, date=date, time=time, amount=amount, volume=fuel_volume, location=Location.Outdoor.name, tender=tender, gas_type=fuel_type, pump_num=pump_num, indoor_prepay=False, reference_num=None, price=price)
    carwash_txn = scan_for_carwash_from_line(txn_line, line_offsets, f)
    if carwash_txn:
        outdoor_txn.carwash_txn = carwash_txn
    return outdoor_txn


def scan_for_carwash_from_line(txn_line, line_offsets, f, location=None, tender_scan=False):
    """"""
    carwash_found, carwash_match = False, None
    carwash_offset = 0
    l = None
    while not carwash_found:
        carwash_offset += 1
        f.seek(line_offsets[txn_line + carwash_offset])
        l = f.readline()
        carwash_match = re.match(CARWASH_REGEX, l)
        total_due_match = re.match(TOTAL_DUE_REGEX, l)
        if carwash_match:
            carwash_found = True
        elif total_due_match:
            break
    if carwash_found:
        if l[(carwash_match.start('dollars')-1)] != '-':
            tender = None
            if tender_scan:
                tender_found, tender_match = False, None
                tender_offset = carwash_offset
                while not tender_found:
                    tender_offset += 1
                    f.seek(line_offsets[txn_line + tender_offset])
                    l = f.readline()
                    tender_match = re.match(OUTDOOR_TENDER_TYPE_REGEX, l)
                    if tender_match:
                        tender_found = True
                tender = tender_match.group('tender')
            return CarWashTxn(id=None, date=None, time=None, amount=None, location=location, carwash_type=carwash_match.group('type'), tender=tender)
    return None


def get_gas_transaction_from_line(txn_line, line_offsets, f):
    """
    txn_line + 1: Date and time
    txn_line + 2: Indoor or Outdoor
    if Outdoor:
        txn_line + 3: 'Outdoor tmnl: 1'
        txn_line + 6: '    UNLEADED PUR       52.50'
        txn_line + 8: 'Vol     12.592@     4.169'
        txn_line + 17: 'Card Type: Debit', 'Card Type: MASTERCARD'

    elif txn_line + 3 == 'User Session: [0-9]+':
        skip over this probably
        txn_line + 7: 'Fuel Prepay Ref#2740063 Pump 2'
        seek til 'TOTAL DUE             1.59'
        seek line + 2:
            if it == 'Cash                 18.41', ''
                it's cash


    elif indoor and txn_line + 3 == 'User Session: 6064' and txn_line + 6 == 'Original Fuel Prepay Ref#[0-9]+':
        it's a prepay
        create a mapping for the reference number with that transaction
        seek til TOTAL DUE:
            then + 2 == 'Cash                 18.41', 'Debit Card           33.69', '' <-- Probably cash
    elif indoor and txn_line + 6 == 'Original ..'
        pull it out of the map by reference number
        txn_line + 7: 'FUEL PREPAY\s+\-[0-9]+\.[0-9]+ (amount prepaid)
        txn_line + 8: 'UNLEADED PUR       17.00
        txn_line + 9: 'Ticket #923757      Pump 6'
        txn_line + 10: 'Vol      4.078@     4.169'
    """
    f.seek(line_offsets[txn_line])
    txn_line_offset = 0
    txn_id = re.match(FINALIZED_TRANSACTION_REGEX, f.readline()).group('txn_id')
    txn_line_offset += 1
    d_time_line = f.readline()
    dt_match = re.match(DATE_TIME_REGEX, d_time_line)
    date = '{0}/{1}/{2}'.format(dt_match.group('month'), dt_match.group('day'), dt_match.group('yr'))
    time = '{0}:{1}:{2}'.format(dt_match.group('hr'), dt_match.group('min'), dt_match.group('sec'))
    txn_line_offset += 1
    f.seek(line_offsets[txn_line + txn_line_offset])
    loc_tmnl_line = f.readline()
    loc_tmnl_match = re.match(INDOOR_OUTDOOR_REGEX, loc_tmnl_line)
    if loc_tmnl_match.group('location') == Location.Indoor.name:
        # Seek to the next line to test for user session
        txn_line_offset += 1
        f.seek(line_offsets[txn_line + txn_line_offset])
        u_ses_line = f.readline()
        user_session_match = re.match(USER_SESSION_REGEX, u_ses_line)
        if user_session_match:
            # Seek ahead to see if this is a fuel prepay. If TOTAL DUE is hit, it's not
            fuel_prepay_found, fuel_prepay_ref_match = False, None
            original_fuel_prepay_line_found, ofp_match = False, None
            while not (fuel_prepay_found or original_fuel_prepay_line_found):
                txn_line_offset += 1
                f.seek(line_offsets[txn_line + txn_line_offset])
                l = f.readline()
                fuel_prepay_ref_match = re.match(INITIAL_FUEL_PREPAY_REGEX, l)
                ofp_match = re.match(FINAL_FUEL_PREPAY_REGEX, l)
                if fuel_prepay_ref_match:
                    fuel_prepay_found = True
                elif ofp_match:
                    original_fuel_prepay_line_found = True
                elif re.match(TOTAL_DUE_REGEX, l):
                    break
            if fuel_prepay_found:
                return handle_initial_gas_txn(fuel_prepay_ref_match, txn_line_offset, line_offsets, txn_line, f, txn_id, date, time)
            elif original_fuel_prepay_line_found:
                return handle_final_gas_txn(ofp_match, txn_line_offset, line_offsets, txn_line, f, txn_id, date, time)
            else:
                # It's not a prepay, return False, None
                return False, None
        else:
            # Could be a finalized fuel transaction still
            original_fuel_prepay_line_found, ofp_match = False, None
            while not original_fuel_prepay_line_found:
                txn_line_offset += 1
                f.seek(line_offsets[txn_line + txn_line_offset])
                l = f.readline()
                ofp_match = re.match(FINAL_FUEL_PREPAY_REGEX, l)
                if ofp_match:
                    original_fuel_prepay_line_found = True
                elif re.match(TOTAL_DUE_REGEX, l):
                    break
            if original_fuel_prepay_line_found:
                return handle_final_gas_txn(ofp_match, txn_line_offset, line_offsets, txn_line, f, txn_id, date, time)
            else:
                return False, None
    else:
        # It's outdoor, this is a gas txn
        pump_num = loc_tmnl_match.group('terminal')
        return True, handle_outdoor_gas_txn(txn_id, date, time, pump_num, txn_line_offset, line_offsets, txn_line, f)


def merge_txns(prepay_txn, final_txn):
    """"""
    final_txn.amount = prepay_txn.amount
    final_txn.tender = prepay_txn.tender
    final_txn.pump_num = prepay_txn.pump_num
    final_txn.carwash_txn = prepay_txn.carwash_txn
    return final_txn


def get_gas_transactions_for_day(day_path):
    """"""
    gas_txns = []
    carwash_txns = []
    prepay_map = {}
    skip_ref_set = set()
    with open(day_path, 'r') as day:
        txn_lines, line_offsets = [], []
        offset = day.tell()
        for i, line in enumerate(day):
            line_offsets.append(offset)
            offset += len(line)
            m = re.match(FINALIZED_TRANSACTION_REGEX, line)
            if m:
                txn_lines.append(i)
        day.seek(0)
        for txn in txn_lines:
            is_gas_txn, gas_txn = get_gas_transaction_from_line(txn, line_offsets, day)
            if is_gas_txn:
                if gas_txn.indoor_prepay:
                    # put it in the prepay map
                    prepay_map[gas_txn.reference_num] = gas_txn
                elif gas_txn.location == Location.Indoor.name:
                    if gas_txn.reference_num in skip_ref_set:
                        continue
                    skip_ref_set.add(gas_txn.reference_num)
                    if gas_txn.reference_num not in prepay_map:
                        print '\tMissing txn number: {0}'.format(gas_txn.reference_num)
                        continue
                    prepay_txn = prepay_map[gas_txn.reference_num]
                    del prepay_map[gas_txn.reference_num]
                    gas_txn = merge_txns(prepay_txn, gas_txn)
                    gas_txns.append(gas_txn)
                else:
                    gas_txns.append(gas_txn)
            else:
                # Get the location type
                day.seek(line_offsets[txn + 2])
                l_match = re.match(INDOOR_OUTDOOR_REGEX, day.readline())
                carwash_txn = scan_for_carwash_from_line(txn, line_offsets, day, location=l_match.group('location'),tender_scan=True)
                if carwash_txn:
                    carwash_txns.append(carwash_txn)
    return gas_txns, carwash_txns


def get_column_letter_for_column_number(row, col):
    """
    """
    quot, rem = divmod(col-1, 26)
    return((chr(quot-1 + ord("A")) if quot else "") +
           (chr(rem + ord("A")) + str(row)))


def initialize_worksheet(ws):
    """"""
    index = 2
    bold_font = Font(bold=True)
    label_fill = PatternFill(fill_type='solid', start_color=Color('00EEEEEE'))
    label_cell_style = Style(font=bold_font, fill=label_fill)
    for i, label in enumerate(GAS_CELL_LABELS):
        label_cell = ws.cell('A{0}'.format(i + index))
        label_cell.value = label
        label_cell.style = label_cell_style
    return ws


def main(args):
    """"""
    months_directory_path = './'
    excel_workbook_path = './results.xlsx'
    if len(args) > 2:
        months_directory_path = args[1]
        excel_workbook_path = args[2]

    month_directories = [os.path.join(months_directory_path, p) for p in os.listdir(months_directory_path) if isdir(os.path.join(months_directory_path, p))]
    dirs = get_month_directories_to_analyze(month_directories)
    try:
        wb = load_workbook(excel_workbook_path)
    except InvalidFileException:
        wb = Workbook()
        wb.save(excel_workbook_path)
    for d in dirs:
        day_files = [os.path.join(d, p) for p in os.listdir(d) if os.path.splitext(p)[1] == '.txt']
        date_obj = datetime.strptime(str(d).split('/')[-1], '%Y%m')
        ws_title = date_obj.strftime('%B %Y')
        ws = wb.create_sheet(title=ws_title)
        ws = initialize_worksheet(ws)
        for i, df in enumerate(day_files):
            gas_txns, carwash_txns = get_gas_transactions_for_day(df)
            date_obj = datetime.strptime(str(df).split('/')[-1].split('.txt')[0], '%Y%m%d')
            da = DayAnalyzer(gas_txns, date_obj, i+2, carwash_txns)
            print 'adding gas txns for {0}'.format(date_obj)
            da.add_gas_txns_to_worksheet(ws)
        mark_directory_as_analyzed(d)
        wb.save(excel_workbook_path)


if __name__ == '__main__':
    main(sys.argv)